"""
ValidoAI - Comprehensive Reporting System
==========================================
Generate reports in multiple formats: PDF, Excel, CSV, JSON
Focus on Serbian business compliance and financial reporting
"""

import os
import json
import csv
import logging
from datetime import datetime, timedelta
from typing import Dict, Any, List, Optional
import uuid

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False

try:
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib import colors
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from .config import get_db_config
import psycopg2
import psycopg2.extras

logger = logging.getLogger(__name__)

class ReportGenerator:
    """Comprehensive report generator for Serbian businesses"""

    def __init__(self):
        self.config = get_db_config().get_current_config()
        self.output_dir = os.path.join(os.getcwd(), 'reports')
        os.makedirs(self.output_dir, exist_ok=True)

    def get_db_connection(self):
        """Get database connection"""
        return psycopg2.connect(
            host=self.config['host'],
            port=self.config['port'],
            database=self.config['database'],
            user=self.config['user'],
            password=self.config['password']
        )

    def generate_company_report(self, company_id: str = None, format: str = 'json',
                               start_date: str = None, end_date: str = None) -> str:
        """Generate comprehensive company report"""

        conn = self.get_db_connection()
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

        try:
            # Get company information
            if company_id:
                cur.execute("""
                    SELECT * FROM companies
                    WHERE companies_id = %s AND status = 'active'
                """, (company_id,))
            else:
                cur.execute("SELECT * FROM companies WHERE status = 'active'")

            companies = cur.fetchall()

            report_data = {
                'generated_at': datetime.now().isoformat(),
                'report_type': 'company_report',
                'companies': []
            }

            for company in companies:
                company_data = dict(company)

                # Get company statistics
                cur.execute("""
                    SELECT
                        COUNT(*) as total_invoices,
                        SUM(total_amount) as total_revenue,
                        AVG(total_amount) as avg_invoice_value,
                        COUNT(CASE WHEN payment_status = 'paid' THEN 1 END) as paid_invoices,
                        COUNT(CASE WHEN payment_status = 'pending' THEN 1 END) as pending_invoices,
                        COUNT(CASE WHEN payment_status = 'overdue' THEN 1 END) as overdue_invoices
                    FROM invoices
                    WHERE company_id = %s AND status = 'issued'
                """, (company['companies_id'],))

                stats = cur.fetchone()
                company_data['statistics'] = dict(stats) if stats else {}

                # Get recent invoices
                cur.execute("""
                    SELECT
                        invoice_number, invoice_date, due_date, total_amount,
                        payment_status, currency
                    FROM invoices
                    WHERE company_id = %s AND status = 'issued'
                    ORDER BY invoice_date DESC
                    LIMIT 10
                """, (company['companies_id'],))

                recent_invoices = cur.fetchall()
                company_data['recent_invoices'] = [dict(inv) for inv in recent_invoices]

                report_data['companies'].append(company_data)

            # Generate report in requested format
            return self._generate_report(report_data, f"company_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}", format)

        finally:
            cur.close()
            conn.close()

    def generate_financial_report(self, company_id: str = None, format: str = 'excel',
                                 start_date: str = None, end_date: str = None) -> str:
        """Generate comprehensive financial report"""

        conn = self.get_db_connection()
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

        try:
            # Set date range
            if not start_date:
                start_date = (datetime.now() - timedelta(days=365)).strftime('%Y-%m-%d')
            if not end_date:
                end_date = datetime.now().strftime('%Y-%m-%d')

            where_clause = "AND invoice_date BETWEEN %s AND %s"
            params = [start_date, end_date]

            if company_id:
                where_clause += " AND company_id = %s"
                params.append(company_id)

            # Get financial summary
            cur.execute(f"""
                SELECT
                    DATE_TRUNC('month', invoice_date) as month,
                    COUNT(*) as invoice_count,
                    SUM(subtotal) as total_subtotal,
                    SUM(pdv_amount) as total_pdv,
                    SUM(total_amount) as total_revenue,
                    AVG(total_amount) as avg_invoice_value,
                    COUNT(CASE WHEN payment_status = 'paid' THEN 1 END) as paid_invoices,
                    COUNT(CASE WHEN payment_status = 'pending' THEN 1 END) as pending_invoices,
                    COUNT(CASE WHEN payment_status = 'overdue' THEN 1 END) as overdue_invoices
                FROM invoices
                WHERE status = 'issued' {where_clause}
                GROUP BY DATE_TRUNC('month', invoice_date)
                ORDER BY month DESC
            """, params)

            monthly_data = cur.fetchall()

            # Get overall statistics
            cur.execute(f"""
                SELECT
                    COUNT(*) as total_invoices,
                    SUM(subtotal) as total_subtotal,
                    SUM(pdv_amount) as total_pdv,
                    SUM(total_amount) as total_revenue,
                    AVG(total_amount) as avg_invoice_value,
                    COUNT(CASE WHEN payment_status = 'paid' THEN 1 END) as paid_invoices,
                    COUNT(CASE WHEN payment_status = 'overdue' THEN 1 END) as overdue_invoices,
                    MAX(total_amount) as largest_invoice,
                    MIN(total_amount) as smallest_invoice
                FROM invoices
                WHERE status = 'issued' {where_clause}
            """, params)

            overall_stats = cur.fetchone()

            report_data = {
                'generated_at': datetime.now().isoformat(),
                'report_type': 'financial_report',
                'period': {
                    'start_date': start_date,
                    'end_date': end_date
                },
                'overall_statistics': dict(overall_stats) if overall_stats else {},
                'monthly_breakdown': [dict(row) for row in monthly_data],
                'company_filter': company_id
            }

            return self._generate_report(report_data, f"financial_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}", format)

        finally:
            cur.close()
            conn.close()

    def generate_tax_report(self, company_id: str = None, format: str = 'pdf',
                           year: int = None) -> str:
        """Generate Serbian PDV (VAT) tax report"""

        if not year:
            year = datetime.now().year

        conn = self.get_db_connection()
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

        try:
            where_clause = "AND EXTRACT(YEAR FROM invoice_date) = %s"
            params = [year]

            if company_id:
                where_clause += " AND company_id = %s"
                params.append(company_id)

            # Get tax data by month
            cur.execute(f"""
                SELECT
                    DATE_TRUNC('month', invoice_date) as month,
                    SUM(subtotal) as taxable_base,
                    SUM(pdv_amount) as pdv_amount,
                    SUM(total_amount) as total_with_pdv,
                    COUNT(*) as invoice_count
                FROM invoices
                WHERE status = 'issued' AND pdv_rate > 0 {where_clause}
                GROUP BY DATE_TRUNC('month', invoice_date)
                ORDER BY month
            """, params)

            tax_data = cur.fetchall()

            # Calculate quarterly totals
            quarterly_totals = []
            for quarter in range(1, 5):
                quarter_start = f"{year}-{((quarter-1)*3)+1:02d}-01"
                quarter_end = f"{year}-{quarter*3:02d}-{31 if quarter*3 != 12 else 31}"

                quarterly_data = [row for row in tax_data if
                                quarter_start <= row['month'].strftime('%Y-%m-%d')[:10] <= quarter_end]

                if quarterly_data:
                    total_base = sum(row['taxable_base'] for row in quarterly_data)
                    total_pdv = sum(row['pdv_amount'] for row in quarterly_data)
                    quarterly_totals.append({
                        'quarter': quarter,
                        'taxable_base': total_base,
                        'pdv_amount': total_pdv,
                        'total_with_pdv': total_base + total_pdv
                    })

            report_data = {
                'generated_at': datetime.now().isoformat(),
                'report_type': 'tax_report',
                'year': year,
                'company_filter': company_id,
                'monthly_data': [dict(row) for row in tax_data],
                'quarterly_totals': quarterly_totals,
                'annual_total': {
                    'taxable_base': sum(row['taxable_base'] for row in tax_data),
                    'pdv_amount': sum(row['pdv_amount'] for row in tax_data),
                    'total_with_pdv': sum(row['total_with_pdv'] for row in tax_data)
                }
            }

            return self._generate_report(report_data, f"tax_report_{year}_{datetime.now().strftime('%Y%m%d_%H%M%S')}", format)

        finally:
            cur.close()
            conn.close()

    def generate_customer_report(self, format: str = 'csv') -> str:
        """Generate customer analysis report"""

        conn = self.get_db_connection()
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

        try:
            cur.execute("""
                SELECT
                    c.*,
                    COUNT(i.invoices_id) as total_invoices,
                    SUM(i.total_amount) as total_revenue,
                    AVG(i.total_amount) as avg_invoice_value,
                    MAX(i.invoice_date) as last_invoice_date,
                    COUNT(CASE WHEN i.payment_status = 'paid' THEN 1 END) as paid_invoices,
                    COUNT(CASE WHEN i.payment_status = 'overdue' THEN 1 END) as overdue_invoices
                FROM customers c
                LEFT JOIN invoices i ON c.customers_id = i.customer_id
                WHERE c.status = 'active'
                GROUP BY c.customers_id
                ORDER BY total_revenue DESC
            """)

            customers = cur.fetchall()

            report_data = {
                'generated_at': datetime.now().isoformat(),
                'report_type': 'customer_report',
                'customers': [dict(customer) for customer in customers]
            }

            return self._generate_report(report_data, f"customer_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}", format)

        finally:
            cur.close()
            conn.close()

    def _generate_report(self, data: Dict[str, Any], filename: str, format: str) -> str:
        """Generate report in specified format"""

        if format.lower() == 'json':
            return self._generate_json_report(data, filename)
        elif format.lower() == 'csv':
            return self._generate_csv_report(data, filename)
        elif format.lower() == 'excel':
            return self._generate_excel_report(data, filename)
        elif format.lower() == 'pdf':
            return self._generate_pdf_report(data, filename)
        else:
            raise ValueError(f"Unsupported format: {format}")

    def _generate_json_report(self, data: Dict[str, Any], filename: str) -> str:
        """Generate JSON report"""
        filepath = os.path.join(self.output_dir, f"{filename}.json")

        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, ensure_ascii=False, default=str)

        logger.info(f"✅ JSON report generated: {filepath}")
        return filepath

    def _generate_csv_report(self, data: Dict[str, Any], filename: str) -> str:
        """Generate CSV report"""
        filepath = os.path.join(self.output_dir, f"{filename}.csv")

        # Flatten the data structure for CSV
        flattened_data = self._flatten_report_data(data)

        if flattened_data:
            with open(filepath, 'w', newline='', encoding='utf-8') as f:
                writer = csv.DictWriter(f, fieldnames=flattened_data[0].keys())
                writer.writeheader()
                writer.writerows(flattened_data)

        logger.info(f"✅ CSV report generated: {filepath}")
        return filepath

    def _generate_excel_report(self, data: Dict[str, Any], filename: str) -> str:
        """Generate Excel report"""
        if not OPENPYXL_AVAILABLE:
            logger.warning("OpenPyXL not available, falling back to CSV")
            return self._generate_csv_report(data, filename)

        filepath = os.path.join(self.output_dir, f"{filename}.xlsx")

        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Report"

        # Add headers and data based on report type
        self._populate_excel_worksheet(worksheet, data)

        workbook.save(filepath)
        logger.info(f"✅ Excel report generated: {filepath}")
        return filepath

    def _generate_pdf_report(self, data: Dict[str, Any], filename: str) -> str:
        """Generate PDF report"""
        if not REPORTLAB_AVAILABLE:
            logger.warning("ReportLab not available, falling back to JSON")
            return self._generate_json_report(data, filename)

        filepath = os.path.join(self.output_dir, f"{filename}.pdf")

        doc = SimpleDocTemplate(filepath, pagesize=A4)
        story = []

        # Add title
        styles = getSampleStyleSheet()
        title = Paragraph(f"ValidoAI Report - {data.get('report_type', 'Unknown')}", styles['Title'])
        story.append(title)
        story.append(Spacer(1, 20))

        # Add generation info
        info_text = f"Generated: {data.get('generated_at', 'Unknown')}"
        story.append(Paragraph(info_text, styles['Normal']))
        story.append(Spacer(1, 20))

        # Add report-specific content
        self._populate_pdf_content(story, data, styles)

        doc.build(story)
        logger.info(f"✅ PDF report generated: {filepath}")
        return filepath

    def _flatten_report_data(self, data: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Flatten nested report data for CSV export"""
        report_type = data.get('report_type', 'unknown')

        if report_type == 'company_report':
            flattened = []
            for company in data.get('companies', []):
                base_row = {
                    'company_name': company.get('company_name', ''),
                    'tax_id': company.get('tax_id', ''),
                    'total_invoices': company.get('statistics', {}).get('total_invoices', 0),
                    'total_revenue': company.get('statistics', {}).get('total_revenue', 0),
                }
                flattened.append(base_row)
            return flattened

        elif report_type == 'financial_report':
            return data.get('monthly_breakdown', [])

        elif report_type == 'customer_report':
            return data.get('customers', [])

        else:
            # Generic flattening
            return [data]

    def _populate_excel_worksheet(self, worksheet, data: Dict[str, Any]):
        """Populate Excel worksheet with report data"""
        report_type = data.get('report_type', 'unknown')

        # Add headers
        headers = ['Field', 'Value']
        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

        # Add data based on report type
        row_num = 2

        if report_type == 'financial_report':
            # Add overall statistics
            for key, value in data.get('overall_statistics', {}).items():
                worksheet.cell(row=row_num, column=1).value = key
                worksheet.cell(row=row_num, column=2).value = value
                row_num += 1

            # Add monthly breakdown
            row_num += 2
            worksheet.cell(row=row_num, column=1).value = "Monthly Breakdown"
            row_num += 1

            monthly_headers = ['Month', 'Invoice Count', 'Total Revenue', 'Paid Invoices']
            for col_num, header in enumerate(monthly_headers, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = header
                cell.font = Font(bold=True)

            row_num += 1

            for month_data in data.get('monthly_breakdown', []):
                worksheet.cell(row=row_num, column=1).value = month_data.get('month')
                worksheet.cell(row=row_num, column=2).value = month_data.get('invoice_count')
                worksheet.cell(row=row_num, column=3).value = month_data.get('total_revenue')
                worksheet.cell(row=row_num, column=4).value = month_data.get('paid_invoices')
                row_num += 1

    def _populate_pdf_content(self, story, data: Dict[str, Any], styles):
        """Populate PDF content based on report type"""
        report_type = data.get('report_type', 'unknown')

        if report_type == 'tax_report':
            # Add tax report content
            story.append(Paragraph("PDV Tax Report", styles['Heading2']))
            story.append(Spacer(1, 12))

            # Add annual total
            annual_total = data.get('annual_total', {})
            story.append(Paragraph(f"Year: {data.get('year', 'Unknown')}", styles['Normal']))
            story.append(Paragraph(f"Taxable Base: {annual_total.get('taxable_base', 0):,.2f} RSD", styles['Normal']))
            story.append(Paragraph(f"PDV Amount: {annual_total.get('pdv_amount', 0):,.2f} RSD", styles['Normal']))
            story.append(Paragraph(f"Total with PDV: {annual_total.get('total_with_pdv', 0):,.2f} RSD", styles['Normal']))

        elif report_type == 'financial_report':
            # Add financial report content
            story.append(Paragraph("Financial Report", styles['Heading2']))
            story.append(Spacer(1, 12))

            stats = data.get('overall_statistics', {})
            story.append(Paragraph(f"Total Invoices: {stats.get('total_invoices', 0)}", styles['Normal']))
            story.append(Paragraph(f"Total Revenue: {stats.get('total_revenue', 0):,.2f} RSD", styles['Normal']))
            story.append(Paragraph(f"Average Invoice: {stats.get('avg_invoice_value', 0):,.2f} RSD", styles['Normal']))

    def list_available_reports(self) -> List[str]:
        """List all generated reports"""
        if os.path.exists(self.output_dir):
            return [f for f in os.listdir(self.output_dir) if f.endswith(('.json', '.csv', '.xlsx', '.pdf'))]
        return []

    def cleanup_old_reports(self, days: int = 30):
        """Clean up reports older than specified days"""
        cutoff_date = datetime.now() - timedelta(days=days)

        for filename in self.list_available_reports():
            filepath = os.path.join(self.output_dir, filename)
            if os.path.getctime(filepath) < cutoff_date.timestamp():
                os.remove(filepath)
                logger.info(f"Cleaned up old report: {filename}")

# Global report generator instance
report_generator = ReportGenerator()

# Utility functions
def generate_company_report(company_id: str = None, format: str = 'json') -> str:
    """Generate company report"""
    return report_generator.generate_company_report(company_id, format)

def generate_financial_report(company_id: str = None, format: str = 'excel',
                             start_date: str = None, end_date: str = None) -> str:
    """Generate financial report"""
    return report_generator.generate_financial_report(company_id, format, start_date, end_date)

def generate_tax_report(company_id: str = None, format: str = 'pdf', year: int = None) -> str:
    """Generate tax report"""
    return report_generator.generate_tax_report(company_id, format, year)

def generate_customer_report(format: str = 'csv') -> str:
    """Generate customer report"""
    return report_generator.generate_customer_report(format)

# Export key components
__all__ = [
    'ReportGenerator',
    'report_generator',
    'generate_company_report',
    'generate_financial_report',
    'generate_tax_report',
    'generate_customer_report'
]

